from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from django.http import HttpResponse
from jobscanner.models import Recrutier, Attendee, ScanLog

from io import BytesIO
import zipfile
import qrcode
from PIL import Image as PILImage
import openpyxl
import random

# Create your views here.
PAGE_SIZE = 10

# Helper Functions
def get_hostname(request):
    return f"{request.scheme}://{request.get_host()}"


# def qr_generator(host_name: str, freelancer: Attendee):
#     url = f"{host_name}/profile/{freelancer.pk}"
#     qr = qrcode.make(url)
#     buffer = BytesIO()
#     qr.save(buffer, format='PNG')
#     buffer.seek(0)

#     return buffer


def qr_generator_svg(host_name: str, attendee: Attendee):
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_L)
    qr.add_data(f"{host_name}/profile/{attendee.pk}")
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    # Save the image to a buffer with 300 DPI
    img_buffer = BytesIO()
    qr_image.save(img_buffer, format="PNG", dpi=(300, 300))
    img_buffer.seek(0)
    return img_buffer

def get_login_code(code_list):
    code = random.randint(10000, 50000)
    while code in code_list:
        code = random.randint(10000, 50000)
    
    return code

# Home Page [Login_or_Scan]
def index(request):
    template = "index.html"
    ctx = {}
    if "is_authenticated" in request.session:
        template = "scanner.html"

    if request.method == "POST":
        login_code = request.POST.get("code", "")
        recrutier = Recrutier.objects.filter(code=login_code).first()
        if recrutier:
            request.session['is_authenticated'] = True
            request.session['recrutier_pk'] = recrutier.pk
            return redirect(reverse("home"))
        else:
            ctx['message'] = "Invalid Code!"
    
    return render(request, template, ctx)



# Profile Display [Display profile | Back Home]
def profile(request, pk):
    if "is_authenticated" in request.session:
        freelancer_profile = get_object_or_404(Attendee, pk=pk)
        recrutier = get_object_or_404(Recrutier, pk=request.session['recrutier_pk'])
        scanlog, created = ScanLog.objects.get_or_create(attendee=freelancer_profile, recrutier=recrutier)
        if created:
            freelancer_profile.visits += 1
            recrutier.scanned_counts += 1
            recrutier.save()
            freelancer_profile.save()

        return render(request, "profile.html", {
            "profile" : freelancer_profile,
            'comment' : scanlog.comment
        })
    
    return redirect(reverse("home"))

def comment(request, pk):
   if "is_authenticated" in request.session:
        freelancer_profile = get_object_or_404(Attendee, pk=pk)
        recrutier = get_object_or_404(Recrutier, pk=request.session['recrutier_pk'])
        scan_log, _ = ScanLog.objects.get_or_create(attendee=freelancer_profile, recrutier=recrutier)
        scan_log.comment = request.POST.get("comment", "")
        scan_log.save()

        if _:
            freelancer_profile.visits += 1
            recrutier.scanned_counts += 1
            recrutier.save()
            freelancer_profile.save()

        return redirect(reverse("profile", kwargs={"pk" : pk}))


def scanned(request):
    """
        Filter Scan Logs to each recrutier
        Then split into pages
        Finally return list of comments and freelancers
    """
    if "recrutier_pk" in request.session:
        recrutier = get_object_or_404(Recrutier, pk=request.session['recrutier_pk'])
        scan_log = ScanLog.objects.filter(recrutier=recrutier).order_by("pk")
        paginator = Paginator(scan_log, PAGE_SIZE)
        page_number = request.GET.get('page', "1")
        page_obj = paginator.get_page(page_number)

        return render(request, "scanned_table.html", {
            "page_obj" : page_obj
        })
    return redirect(reverse("home"))


# TODO
def dashboard(request):
    # Display a table with pagination, you can refer to scanned function and scanned_table.html
    # List 10 per page preferable to use PAGE_SIZE
    if request.user.is_authenticated:
        recrutier = Recrutier.objects.all().order_by("name")
        paginator = Paginator(recrutier, PAGE_SIZE)
        page_number = request.GET.get('page', "1")
        page_obj = paginator.get_page(page_number)

        return render(request, "admin_dashboard.html", {
            "page_obj" : page_obj
        })
    return redirect(reverse("home"))

# TODO
def detailed_dashboard(request, pk):
    # Here list all scanned freelancers as you wish as card or in table as you like
    # if you do pagination make sure do not make more than 10 per page and preferable to use PAGE_SIZE
    if request.user.is_authenticated:
        recrutier = get_object_or_404(Recrutier, pk=pk)
        return render(request, "profile_cards.html", {
            "scanlogs" : recrutier.scanned_logs.all()
        })
    return redirect(reverse("home"))

# TODO
def upload_freelancers(request):
    # Take csv file From Zidan [name, email, phone_number, location, track, job_interest, cv_link]
    # Insert in Database using create of Attendee Class
    # then generate csv file [name, email, phone, qr_code]
    # qr_code function that takes host_name and freelancer_obj to build dynamic link, is written for you, it generates buffer to use and create image then embeded it in xlsx
    if request.method == 'POST' and 'file' in request.FILES:
        # 1. Read the uploaded Excel file
        uploaded_file = request.FILES['file']
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        host_name = get_hostname(request)
        
        # 2. Extract data and insert into the Attendee table
        headers = [cell.value for cell in sheet[1]]  # Read headers from the first row
        data_rows = sheet.iter_rows(min_row=2, values_only=True)  # Read data rows
        
        for row in data_rows:
            row_data = dict(zip(headers, row))
            # Map Excel columns to Attendee fields
            try:
                freelancer = Attendee(
                    name=row_data.get("name"),
                    email=row_data.get("email"),
                    phone_number=row_data.get("phone_number"),
                    linkedin=row_data.get("linkedin"),
                    track=row_data.get("track"),
                    location=row_data.get("location") or "",
                    age=row_data.get("age") or 1,
                    cv_url=row_data.get("cv_link"),
                )
                freelancer.save()
            except:
                pass

        # Create an in-memory ZIP file
        zip_buffer = BytesIO()
        attendees = Attendee.objects.all()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for attendee in attendees:
                # Generate the QR code (assuming qr_generator returns a BytesIO object)
                qr_buffer = qr_generator_svg(host_name, attendee)

                # Name the image file with the attendee's name and email
                file_name = f"{attendee.email}.png"

                # Add the image to the ZIP file
                zip_file.writestr(file_name, qr_buffer.getvalue())

        # Prepare the ZIP file for download
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="attendee_qrcodes.zip"'

        return response
    elif request.method == "GET":
        return render(request, "upload_freelancers.html")
    
def upload_recrutiers(request):
    if request.method == 'POST' and 'file' in request.FILES:
        # 1. Read the uploaded Excel file
        uploaded_file = request.FILES['file']
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        codes = []
        
        # 2. Extract data and insert into the Attendee table
        headers = [cell.value for cell in sheet[1]]  # Read headers from the first row
        data_rows = sheet.iter_rows(min_row=2, values_only=True)  # Read data rows
        
        for row in data_rows:
            row_data = dict(zip(headers, row))
            recrutier = Recrutier.objects.filter(name=row_data.get("company_name"))
            if not recrutier:
                # Map Excel columns to Attendee fields
                generated_login_code = get_login_code(codes)
                codes.append(generated_login_code)
                recrutier = Recrutier(
                    name=row_data.get("company_name"),
                    rep_name=row_data.get("representive_name"),
                    job_title=row_data.get("job_title"),
                    code=generated_login_code
                )
                recrutier.save()



        # 3. Generate a new Excel file with QR codes
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Recrutiers with Login Codes"

        # Add headers to the new sheet
        headers = ["Company Name", "Login Code"]
        new_sheet.append(headers)

        # Fetch all freelancers and generate QR codes
        recrutiers = Recrutier.objects.all()
        for recrutier in recrutiers:
            # Call the qr_generator function (assume it returns a BytesIO object)
            
            # Write freelancer data into the new sheet
            new_row = [recrutier.name, recrutier.code]
            new_sheet.append(new_row)

        # 4. Return the new Excel file as a response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Recrutiers_with_Login_Codes.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    elif request.method == "GET":
        return render(request, "upload_freelancers.html")
    
def download_leads(request):
    if request.user.is_authenticated:
        recrutiers = Recrutier.objects.all()
        new_wb = openpyxl.Workbook()
        for recrutier in recrutiers:
            scan_logs = recrutier.scanned_logs.all()
            new_sheet = new_wb.create_sheet(f"{recrutier.name}")
            headers = ["Name", "Email", "Phone", "ALX Track", "Comment"]
            new_sheet.append(headers)

            for scan_log in scan_logs:
                attendee = scan_log.attendee
                new_sheet.append([
                    attendee.name,
                    attendee.email,
                    attendee.phone_number,
                    attendee.track,
                    scan_log.comment,
                ])

        # 4. Return the new Excel file as a response
        
        new_wb.remove(new_wb["Sheet"])
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Leads.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    return redirect(reverse("home"))

def download_attendees(request):
    if "recrutier_pk" in request.session:
        new_wb = openpyxl.Workbook()
        recrutier = Recrutier.objects.get(pk=request.session['recrutier_pk'])
        scan_logs = recrutier.scanned_logs.all()
        new_sheet = new_wb.active
        headers = ["Name", "Email", "Age", "Phone",  "Comment", "Location", "Linkedin",  "CV Link" , "ALX Track"]
        new_sheet.append(headers)

        for scan_log in scan_logs:
            attendee = scan_log.attendee
            new_sheet.append([
                attendee.name,
                attendee.email,
                attendee.age,
                attendee.phone_number,
                scan_log.comment,
                attendee.location,
                attendee.linkedin,
                attendee.cv_url,
                attendee.track,
            ])

        # 4. Return the new Excel file as a response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Attendees.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    return redirect(reverse("home"))

def qr_code_download(request, pk):
    attendee = Attendee.objects.get(pk=pk)
    host_name = get_hostname(request)
    qr_buffer = qr_generator_svg(host_name, attendee)
    # Name the image file with the attendee's name and email
    file_name = f"{attendee.email}.png"
    qr_buffer.seek(0)
    response = HttpResponse(qr_buffer, content_type="image/png")
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response


def download_unvisited(request):
    if request.user.is_authenticated:
        unvisited_attendee = Attendee.objects.filter(visits=0)
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Recrutiers with Login Codes"

        # Add headers to the new sheet
        headers = ["Email", "Name"]
        new_sheet.append(headers)
        for attendee in unvisited_attendee:
            # Write freelancer data into the new sheet
            new_row = [attendee.email, attendee.name]
            new_sheet.append(new_row)

        # 4. Return the new Excel file as a response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Unvisited_Attendees.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response